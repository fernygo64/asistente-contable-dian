import io
import openpyxl


def test_columna_fantasma_al_final_se_descarta():
    """
    Bug real: un archivo Excel puede tener una columna que nunca tuvo
    título ni dato, pero conserva formato residual (de una columna que
    se borró) — Excel extiende su "rango usado" con esa columna, y el
    detector la tomaba como una columna 124 real (vacía, sin título),
    lo que dejaba un separador de más al final de cada fila exportada.
    """
    from app.services.plantilla_inferencia_service import detectar_estructura_archivo_plano

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "EMPRESA SAS"
    titulos = ["TIPO DE COMPROBANTE (OBLIGATORIO)", "CÓDIGO COMPROBANTE  (OBLIGATORIO)", "CUENTA CONTABLE   (OBLIGATORIO)"]
    for i, t in enumerate(titulos, start=1):
        ws.cell(row=5, column=i, value=t)
    ws.cell(row=6, column=1, value="G")
    ws.cell(row=6, column=2, value=1)
    ws.cell(row=6, column=3, value=5195250000)
    ws.cell(row=6, column=4).font = openpyxl.styles.Font(bold=True)  # columna fantasma: solo formato, sin título ni dato

    buf = io.BytesIO()
    wb.save(buf)

    r = detectar_estructura_archivo_plano(buf.getvalue())
    assert len(r["columnas"]) == 3
    assert [c["label"] for c in r["columnas"]] == titulos
