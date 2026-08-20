import io
import openpyxl

from tests.test_documentos import _zip_bytes, _xml


def test_exportacion_completa_desde_archivo_real_de_siigo(client, empresa_a):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "EMPRESA DE PRUEBA SAS"
    ws["A2"] = "MODELO PARA LA IMPORTACION DE MOVIMIENTO CONTABLE - MODELO GENERAL"
    ws["A3"] = "De :  ENE  1/2026   A :  JUL 31/2026"
    encabezados = [
        "TIPO DE COMPROBANTE (OBLIGATORIO)", "CÓDIGO COMPROBANTE  (OBLIGATORIO)", "NÚMERO DE DOCUMENTO",
        "CUENTA CONTABLE   (OBLIGATORIO)", "DÉBITO O CRÉDITO (OBLIGATORIO)", "VALOR DE LA SECUENCIA   (OBLIGATORIO)",
        "AÑO DEL DOCUMENTO", "MES DEL DOCUMENTO", "DÍA DEL DOCUMENTO", "CÓDIGO DEL VENDEDOR",
        "CÓDIGO DE LA CIUDAD", "CÓDIGO DE LA ZONA", "SECUENCIA", "CENTRO DE COSTO", "SUBCENTRO DE COSTO",
        "NIT", "SUCURSAL", "DESCRIPCIÓN DE LA SECUENCIA",
        "NÚMERO DE CHEQUE", "COMPROBANTE ANULADO",
    ]
    for i, titulo in enumerate(encabezados, start=1):
        ws.cell(row=5, column=i, value=titulo)
    ws.cell(row=6, column=19, value=0)
    ws.cell(row=6, column=20, value="N")

    buf = io.BytesIO()
    wb.save(buf)

    r_detectar = client.post(f"/empresas/{empresa_a['id']}/plantillas/inferir-desde-ejemplo",
                              files={"archivo": ("ejemplo.xlsx", buf.getvalue(),
                                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r_detectar.status_code == 200, r_detectar.text
    estructura = r_detectar.json()
    assert len(estructura["columnas"]) == 20

    origenes = {c["label"].strip(): c["source"] for c in estructura["columnas"]}
    assert origenes["TIPO DE COMPROBANTE (OBLIGATORIO)"] == "tipo_comprobante"
    assert origenes["CÓDIGO DEL VENDEDOR"] == "fijo"
    assert origenes["SECUENCIA"] == "secuencia_linea"
    assert origenes["CENTRO DE COSTO"] == "centro_costo"
    assert origenes["NÚMERO DE CHEQUE"] == "fijo"
    assert origenes["CÓDIGO COMPROBANTE  (OBLIGATORIO)"] == "fijo"
    assert origenes["CÓDIGO DE LA CIUDAD"] == "fijo"

    client.patch(f"/empresas/{empresa_a['id']}/comprobantes-por-tipo", json={"comprobante_factura_recibida": "G"})
    client.patch(f"/empresas/{empresa_a['id']}/cuentas-base", json={"cuenta_proveedores": "220501"})
    client.post(f"/empresas/{empresa_a['id']}/cuentas", json={"codigo": "513595", "nombre": "Gastos"})

    zip_contenido = _zip_bytes({"F.xml": _xml("F900", "cufe-siigo-completo", "900980900",
                                               subtotal="50000", total="50000")})
    client.post(f"/empresas/{empresa_a['id']}/documentos/cargar",
                files=[("documentos", ("d.zip", zip_contenido, "application/zip"))])
    factura = client.get(f"/empresas/{empresa_a['id']}/documentos", params={"nit_emisor": "900980900"}).json()[0]
    client.post(f"/empresas/{empresa_a['id']}/documentos/{factura['id']}/partida/generar",
                json={"cuenta_gasto_codigo": "513595", "contrapartida": "proveedores"})

    r_plantilla = client.post(f"/empresas/{empresa_a['id']}/plantillas", json={
        "nombre": "Siigo Pyme completa", "sistema_contable": "siigo_pyme",
        "columnas": estructura["columnas"], "delimitador": estructura["delimitador"],
    })
    assert r_plantilla.status_code == 201, r_plantilla.text
    plantilla_id = r_plantilla.json()["id"]

    r_export = client.post(f"/empresas/{empresa_a['id']}/exportaciones/generar",
                            json={"plantilla_id": plantilla_id, "factura_ids": [factura["id"]]})
    assert r_export.status_code == 200, r_export.text
    lineas = r_export.content.decode("utf-8").strip().split("\r\n")
    encabezado_out = lineas[0].split("|")
    fila1 = lineas[1].split("|")
    valores = dict(zip(encabezado_out, fila1))

    assert valores["TIPO DE COMPROBANTE (OBLIGATORIO)"] == "G"
    assert valores["CÓDIGO DEL VENDEDOR"] == "1"
    assert valores["SECUENCIA"] == "1"
    assert valores["CENTRO DE COSTO"] == "0"
    assert valores["NÚMERO DE CHEQUE"] == "0"
    assert valores["COMPROBANTE ANULADO"] == "N"
    assert valores["CÓDIGO COMPROBANTE  (OBLIGATORIO)"] == ""
    assert valores["CÓDIGO DE LA CIUDAD"] == ""
