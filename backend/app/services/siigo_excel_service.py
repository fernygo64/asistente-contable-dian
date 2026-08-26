"""Generación del Modelo General SIIGO Pyme en XLSX, listo desde A1:DS."""
from __future__ import annotations

from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.services import export_service
from app.services.siigo_pyme_extendido import reprocesar_columnas_siigo

TEMPLATE = Path(__file__).resolve().parents[1] / "resources" / "siigo_modelo_general_base_v5.xlsx"

# Columnas que conviene escribir como números reales de Excel.
_NUMERICAS = {5, 6, 7, 8, 9, 13, 14, 15, 17, 19, 21, 22, 23, 24, 25, 26, 27,
             31, 33, 34, 35, 36, 38, 39, 41, 42, 43, 44, 46, 47, 48, 49,
             51, 52, 53, 54, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 67, 68,
             69, 70, 72, 73, 74, 76, 77, 79, 80, 81, 82, 83, 84, 85, 89, 90,
             93, 94, 95, 97, 98, 99, 100, 103, 104, 105, 106, 107, 108, 118,
             119, 120, 121, 123}


def _coerce(valor, col_idx):
    if valor is None:
        return ""
    # Cuenta, NIT, comprobantes, producto y textos deben conservar ceros/espacios.
    if col_idx in _NUMERICAS:
        t = str(valor).strip()
        if t == "":
            return valor
        try:
            n = float(t.replace(",", "."))
            return int(n) if n.is_integer() else n
        except Exception:
            return valor
    return valor



def columnas_modelo_general():
    wb = load_workbook(TEMPLATE, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    columnas = [{"label": ws.cell(5, c).value or "", "source": "fijo", "valor_fijo": ""} for c in range(1, 124)]
    wb.close()
    return reprocesar_columnas_siigo(columnas)

def generar_xlsx(db, empresa, plantilla, facturas, numeros_documento=None):
    columnas, filas, total = export_service.generar_filas(
        db, empresa, plantilla, facturas, numeros_documento=numeros_documento
    )
    if len(columnas) != 123:
        raise ValueError(f"El Modelo General SIIGO debe tener 123 columnas; la plantilla tiene {len(columnas)}.")

    wb = load_workbook(TEMPLATE)
    ws = wb[wb.sheetnames[0]]
    encabezados_modelo = [ws.cell(5, c).value or "" for c in range(1, 124)]
    encabezados = [c.get("label", "") for c in columnas]
    if encabezados != encabezados_modelo:
        raise ValueError("Los 123 encabezados de la plantilla no coinciden exactamente con el Modelo General SIIGO.")

    ws["A1"] = empresa.nombre
    ws["A2"] = "MODELO PARA LA IMPORTACION DE MOVIMIENTO CONTABLE - MODELO GENERAL"
    ws["A3"] = f"FECHA DE GENERACIÓN: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    # La fila 4 del modelo está combinada A4:DS4. Conservar la combinación
    # y limpiar únicamente la celda superior izquierda evita escribir sobre
    # objetos MergedCell de solo lectura.
    ws["A4"] = None

    # Conserva el formato de la primera fila de datos del modelo antes de
    # limpiar movimientos residuales. Así cada fila nueva mantiene el aspecto
    # y tipos visuales del Modelo General, sin arrastrar información histórica.
    estilos_fila = []
    for c in range(1, 124):
        celda = ws.cell(6, c)
        estilos_fila.append({
            "font": copy(celda.font), "fill": copy(celda.fill), "border": copy(celda.border),
            "alignment": copy(celda.alignment), "protection": copy(celda.protection),
            "number_format": celda.number_format,
        })
    altura_fila = ws.row_dimensions[6].height

    # Limpia cualquier movimiento residual de la plantilla.
    if ws.max_row > 5:
        ws.delete_rows(6, ws.max_row - 5)

    for r_idx, valores in enumerate(filas, start=6):
        if altura_fila is not None:
            ws.row_dimensions[r_idx].height = altura_fila
        for c_idx, valor in enumerate(valores, start=1):
            destino = ws.cell(r_idx, c_idx)
            estilo = estilos_fila[c_idx - 1]
            destino.font = copy(estilo["font"]); destino.fill = copy(estilo["fill"]); destino.border = copy(estilo["border"])
            destino.alignment = copy(estilo["alignment"]); destino.protection = copy(estilo["protection"])
            destino.number_format = estilo["number_format"]
            destino.value = _coerce(valor, c_idx)

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), total
